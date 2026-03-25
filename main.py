import flet as ft
import os
import shutil
import urllib.request
import tempfile
import pandas as pd
from fpdf import FPDF
from datetime import datetime, timedelta
from supabase import create_client, Client
from openpyxl import Workbook

# --- SUPABASE CONFIGURATION ---
SUPABASE_URL = "https://lbaquqyzbippicbvmcxr.supabase.co"
SUPABASE_KEY = "sb_publishable_qIs62pb-XO17gSwhXubVqg_2ffU7MOl"

try:
    supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
except Exception as e:
    supabase = None
    print(f"Supabase init error: {e}")

MACHINES = [
    "Presse 1", "Presse 2", "Presse 3", "Presse 4", "Presse 5", "Presse 6",
    "Presse 7", "Presse 8", "Presse 9", "Presse 10", "Presse 11",
    "Compresseur d'air", "Aspirateur CNC", "Malaxeur matière", "Machine de colle",
    "Four machine de colle", "CNC profileuse", "CNC profileuse lourd", "Machine de peinture",
    "Four CNC", "Scotcheuse", "Emballeuse", "Cintreuse", "Groupe électrogène",
    "Four produit final", "Pèse matière 1", "Pèse matière 2", "Pèse matière 3",
    "Aspirateur presse", "Aspirateur manuel 1", "Aspirateur manuel 2", "Moule"
]

def setup_db():
    class SupabaseWrapper:
        def execute(self, query, params=None): pass
        def commit(self): pass
        def cursor(self): return self
        def fetchone(self): return [1]
    return SupabaseWrapper()

db_conn = setup_db()

def main(page: ft.Page):
    try:
        page.title = "BRIKS BY OKBA - Service maintenance"
        page.theme_mode = "dark"
        page.scroll = ft.ScrollMode.AUTO
        page.bgcolor = "#0f0f0f"
        page.padding = 10

        page.logged_in = False
        page.view = "LOGIN"

        page.photo_err_path = ""
        page.photo_sol_path = ""
        page.photo_part_path = ""
        page.current_upload_target = ""

        def show_ui_error(e):
            print(f"FLET UI ERROR: {e.data}")
            try:
                page.snack_bar = ft.SnackBar(
                    ft.Text(f"UI ERROR: {str(e.data)}", color="white"),
                    open=True
                )
                page.update()
            except Exception as ex:
                print("FAILED TO SHOW UI ERROR:", ex)

        page.on_error = show_ui_error

        def request_android_permissions():
            try:
                if hasattr(page, "permission_handler"):
                    page.permission_handler.request_permission(ft.PermissionType.STORAGE)
                    page.permission_handler.request_permission(ft.PermissionType.MANAGE_EXTERNAL_STORAGE)
                elif hasattr(page, "request_permission"):
                    page.request_permission(ft.PermissionType.STORAGE)
            except:
                pass

        request_android_permissions()

        def upload_to_supabase(local_path, folder):
            if not local_path or not os.path.exists(local_path):
                return ""
            try:
                user_clean = page.u_id if hasattr(page, 'u_id') else "anon"
                ts = datetime.now().strftime('%Y%m%d_%H%M%S_%f')
                ext = os.path.splitext(local_path)[1]
                if not ext:
                    ext = ".jpg"

                file_name = f"{folder}/{user_clean}_{ts}{ext}"

                with open(local_path, "rb") as f:
                    supabase.storage.from_("maintenance").upload(file_name, f)

                return supabase.storage.from_("maintenance").get_public_url(file_name)
            except Exception as ex:
                print(f"Upload error: {ex}")
                return ""

        def on_file_result(e):
            if e.files:
                if page.current_upload_target == "ERR":
                    page.photo_err_path = e.files[0].path
                    page.snack_bar = ft.SnackBar(ft.Text("✅ Photo Erreur attachée"))
                elif page.current_upload_target == "SOL":
                    page.photo_sol_path = e.files[0].path
                    page.snack_bar = ft.SnackBar(ft.Text("✅ Photo Solution attachée"))
                elif page.current_upload_target == "PART":
                    page.photo_part_path = e.files[0].path
                    page.snack_bar = ft.SnackBar(ft.Text("✅ Photo Pièce attachée"))

                page.snack_bar.open = True
                page.update()

        file_picker = ft.FilePicker(on_result=on_file_result)
        if file_picker not in page.overlay:
            page.overlay.append(file_picker)

        footer_tag = ft.Text("Made by Okba Bennaim", size=10, italic=True, color="grey500")

        header_brand = ft.Column([
            ft.Text("BRIKS BY OKBA", size=32, weight="bold", color="red"),
            ft.Text("SERVICE MAINTENANCE", size=12, color="red", italic=True),
        ], spacing=0, horizontal_alignment=ft.CrossAxisAlignment.CENTER)

        content_area = ft.Container(expand=True)

        def ch_v(v):
            page.view = v
            refresh()

        def safe_screen(*controls):
            return ft.Container(
                content=ft.Column(
                    list(controls),
                    scroll=ft.ScrollMode.AUTO,
                    horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                    expand=True
                ),
                padding=20,
                expand=True
            )

        def refresh():
            try:
                print(f"REFRESH VIEW => {page.view}")

                screen = None

                if not page.logged_in and page.view == "LOGIN":
                    u_in = ft.TextField(label="Utilisateur", width=320)
                    p_in = ft.TextField(label="Mot de passe", password=True, width=320)
                    login_error = ft.Text("Identifiants incorrects", color="red", visible=False)

                    login_btn = ft.ElevatedButton("ENTRER", width=320, bgcolor="red900")
                    loading_ring = ft.ProgressRing(visible=False, color="red")

                    def login(e):
                        login_btn.disabled = True
                        loading_ring.visible = True
                        login_error.visible = False
                        page.update()

                        try:
                            if supabase is None:
                                raise Exception("Supabase non initialisé")

                            res = supabase.table("users").select("*").eq(
                                "username", (u_in.value or "").lower()
                            ).eq("password", p_in.value or "").execute()

                            print("LOGIN RESPONSE:", res.data)

                            if res.data:
                                page.logged_in = True
                                page.u_id = (u_in.value or "").lower()
                                page.display_name = res.data[0].get("full_name", u_in.value or "Utilisateur")
                                page.view = "HOME"
                                refresh()
                                return
                            else:
                                login_error.value = "Identifiants incorrects ou utilisateur introuvable."
                                login_error.visible = True

                        except Exception as ex:
                            print("LOGIN ERROR:", ex)
                            login_error.value = f"Erreur de connexion: {str(ex)}"
                            login_error.visible = True

                        login_btn.disabled = False
                        loading_ring.visible = False
                        page.update()

                    login_btn.on_click = login

                    screen = safe_screen(
                        ft.Container(height=50),
                        header_brand,
                        u_in,
                        p_in,
                        login_error,
                        loading_ring,
                        login_btn,
                        ft.TextButton("Créer un compte (Sign Up)", on_click=lambda _: ch_v("SIGNUP")),
                        footer_tag
                    )

                elif page.view == "SIGNUP":
                    new_user = ft.TextField(label="Nom d'utilisateur (Login)", width=320)
                    new_full = ft.TextField(label="Nom complet (Affichage)", width=320)
                    new_pass = ft.TextField(label="Mot de passe", password=True, width=320)

                    def register(e):
                        if not new_user.value or not new_pass.value:
                            page.snack_bar = ft.SnackBar(ft.Text("Veuillez remplir tous les champs."))
                            page.snack_bar.open = True
                            page.update()
                            return

                        try:
                            supabase.table("users").insert({
                                "username": new_user.value.lower(),
                                "full_name": new_full.value,
                                "password": new_pass.value
                            }).execute()

                            page.snack_bar = ft.SnackBar(ft.Text("Compte créé avec succès ! Connectez-vous."))
                            page.snack_bar.open = True
                            ch_v("LOGIN")
                        except Exception as ex:
                            page.snack_bar = ft.SnackBar(ft.Text(f"Erreur : {ex}"))
                            page.snack_bar.open = True
                            page.update()

                    screen = safe_screen(
                        ft.Container(height=50),
                        header_brand,
                        ft.Text("CRÉER UN COMPTE", weight="bold", size=20),
                        new_user, new_full, new_pass,
                        ft.ElevatedButton("S'INSCRIRE", on_click=register, width=320, bgcolor="blue900"),
                        ft.TextButton("Retour à la connexion", on_click=lambda _: ch_v("LOGIN")),
                        footer_tag
                    )

                elif page.view == "HOME":
                    screen = ft.Container(
                        content=ft.Column([
                            ft.Row([
                                ft.Column([
                                    ft.Text("BRIKS BY OKBA", size=28, weight="bold", color="red"),
                                    ft.Text("SERVICE MAINTENANCE", size=10, color="red", italic=True),
                                ], spacing=0),
                                ft.IconButton(ft.icons.SETTINGS, on_click=lambda _: ch_v("USER"))
                            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),

                            ft.Text(f"Opérateur: {getattr(page, 'display_name', 'Utilisateur')}", italic=True, color="grey"),
                            ft.Divider(color="red"),

                            ft.Column([
                                ft.ElevatedButton("INTERVENTION TECHNIQUE", icon=ft.icons.BUILD_CIRCLE,
                                                  on_click=lambda _: ch_v("INTER"), width=320, height=55, bgcolor="blue900"),
                                ft.ElevatedButton("DEMANDE PIÈCE DE RECHANGE", icon=ft.icons.SHOPPING_CART,
                                                  on_click=lambda _: ch_v("PART_REQ"), width=320, height=55, bgcolor="orange900"),
                                ft.ElevatedButton("GESTION STOCK (INVENTORY)", icon=ft.icons.INVENTORY,
                                                  on_click=lambda _: ch_v("STOCK_MGR"), width=320, height=55, bgcolor="teal900"),
                                ft.ElevatedButton("HISTORIQUE DES RAPPORTS", icon=ft.icons.HISTORY,
                                                  on_click=lambda _: ch_v("HISTORY"), width=320, height=55),
                                ft.ElevatedButton("TRACKING MOULES", icon=ft.icons.RECYCLING,
                                                  on_click=lambda _: ch_v("MOLD"), width=320, height=55),
                                ft.ElevatedButton("CHECKS QUOTIDIENS / HEBDO", icon=ft.icons.CHECKLIST,
                                                  on_click=lambda _: ch_v("ROUTINE"), width=320, height=55, bgcolor="green900"),
                            ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=15),

                            ft.Divider(),
                            footer_tag
                        ],
                        horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                        scroll=ft.ScrollMode.AUTO,
                        expand=True
                        ),
                        padding=20,
                        expand=True
                    )

                elif page.view == "USER":
                    new_name = ft.TextField(label="Nouveau Nom d'affichage", value=getattr(page, "display_name", ""))
                    new_pw = ft.TextField(label="Nouveau Mot de passe", password=True)

                    def update_profile(e):
                        try:
                            payload = {"full_name": new_name.value}
                            if new_pw.value:
                                payload["password"] = new_pw.value

                            supabase.table("users").update(payload).eq("username", page.u_id).execute()
                            page.display_name = new_name.value
                            page.snack_bar = ft.SnackBar(ft.Text("Profil mis à jour !"))
                            page.snack_bar.open = True
                            ch_v("HOME")
                        except Exception as ex:
                            page.snack_bar = ft.SnackBar(ft.Text(f"Erreur: {ex}"))
                            page.snack_bar.open = True
                            page.update()

                    screen = safe_screen(
                        ft.Row([ft.IconButton(ft.icons.ARROW_BACK, on_click=lambda _: ch_v("HOME")), header_brand]),
                        ft.Text("PARAMÈTRES UTILISATEUR", weight="bold"),
                        new_name, new_pw,
                        ft.ElevatedButton("METTRE À JOUR", on_click=update_profile, bgcolor="blue"),
                        ft.ElevatedButton("DÉCONNEXION", icon=ft.icons.LOGOUT,
                                          on_click=lambda _: (setattr(page, "logged_in", False), ch_v("LOGIN")),
                                          bgcolor="red900"),
                        footer_tag
                    )

                elif page.view == "PART_REQ":
                    p_mach = ft.Dropdown(label="Machine Concernée", options=[ft.dropdown.Option(m) for m in MACHINES])
                    p_name = ft.TextField(label="Nom de la pièce / Référence")
                    p_qty = ft.TextField(label="Quantité", value="1", keyboard_type="number")
                    p_urg = ft.Dropdown(label="Degré d'urgence", options=[
                        ft.dropdown.Option("Normal"),
                        ft.dropdown.Option("Urgent"),
                        ft.dropdown.Option("Critique (Arrêt Machine)")
                    ])

                    def save_part_req(e):
                        try:
                            img_url = upload_to_supabase(page.photo_part_path, "parts")
                            supabase.table("part_requests").insert({
                                "machine": p_mach.value, "piece_nom": p_name.value, "qte": p_qty.value,
                                "urgence": p_urg.value, "photo_path": img_url,
                                "dt": datetime.now().strftime("%d/%m/%Y %H:%M"), "user": page.display_name
                            }).execute()
                            page.photo_part_path = ""
                            ch_v("HOME")
                        except Exception as ex:
                            page.snack_bar = ft.SnackBar(ft.Text(f"Erreur: {ex}"))
                            page.snack_bar.open = True
                            page.update()

                    def pick_part_img(e):
                        page.current_upload_target = "PART"
                        file_picker.pick_files()

                    screen = safe_screen(
                        ft.Row([ft.IconButton(ft.icons.ARROW_BACK, on_click=lambda _: ch_v("HOME")), header_brand]),
                        ft.Text("DEMANDE DE PIÈCE DE RECHANGE", weight="bold", color="orange"),
                        p_mach, p_name, p_qty, p_urg,
                        ft.ElevatedButton("PRENDRE PHOTO PIÈCE", icon=ft.icons.CAMERA_ALT, on_click=pick_part_img),
                        ft.ElevatedButton("ENVOYER LA DEMANDE", on_click=save_part_req, bgcolor="orange", width=320),
                        ft.ElevatedButton("HISTORIQUE DEMANDES", icon=ft.icons.LIST_ALT, on_click=lambda _: ch_v("PART_HISTORY"), width=320),
                        footer_tag
                    )

                elif page.view == "PART_HISTORY":
                    reqs = supabase.table("part_requests").select("*").order("id", desc=True).execute().data
                    lv = ft.ListView(expand=True, spacing=10)
                    for r in reqs:
                        lv.controls.append(ft.Container(content=ft.Column([
                            ft.Text(f"REF: PR-2026-{r['id']} | {r['piece_nom']}", weight="bold"),
                            ft.Text(f"Machine: {r['machine']} | Qte: {r['qte']} | Urgence: {r['urgence']}"),
                            ft.IconButton(ft.icons.PICTURE_AS_PDF, icon_color="orange", on_click=lambda e, row=r: export_part_pdf(row))
                        ]), padding=10, border=ft.border.all(1, "orange"), border_radius=10))

                    screen = safe_screen(
                        ft.Row([ft.IconButton(ft.icons.ARROW_BACK, on_click=lambda _: ch_v("PART_REQ")), header_brand]),
                        ft.Text("HISTORIQUE DES DEMANDES PIÈCES"),
                        lv,
                        footer_tag
                    )

                elif page.view == "STOCK_MGR":
                    stock_lv = ft.ListView(expand=True, spacing=5)

                    def build_stock_list(search=""):
                        try:
                            stock_lv.controls.clear()
                            query = supabase.table("inventory").select("*")
                            if search:
                                query = query.or_(f"designation.ilike.%{search}%,ref.ilike.%{search}%")
                            items = query.execute().data

                            for i in items:
                                is_low = i['stock_qty'] <= i['min_qty']
                                stock_lv.controls.append(ft.ListTile(
                                    leading=ft.Icon(ft.icons.SETTINGS_SUGGEST, color="red" if is_low else "teal"),
                                    title=ft.Text(f"{i['designation']} ({i['ref']})", weight="bold"),
                                    subtitle=ft.Text(f"Emplacement: {i['location']} | Cat: {i['category']}"),
                                    trailing=ft.Column([
                                        ft.Text(f"Qté: {i['stock_qty']}", color="red" if is_low else "white", size=16, weight="bold"),
                                        ft.Text("ALERTE" if is_low else "", color="red", size=10)
                                    ], alignment=ft.MainAxisAlignment.CENTER),
                                    on_click=lambda e, item=i: open_stock_dialog(item)
                                ))
                            page.update()
                        except Exception as ex:
                            print("STOCK ERROR:", ex)

                    def open_stock_dialog(item):
                        qty_edit = ft.TextField(label="Ajuster Quantité (+/-)", value="0", width=100)

                        def update_qty(e):
                            try:
                                new_val = item['stock_qty'] + int(qty_edit.value)
                                supabase.table("inventory").update({"stock_qty": new_val}).eq("id", item['id']).execute()
                                supabase.table("inventory_logs").insert({
                                    "part_id": item['id'], "action": "Manual Adjust", "qty": int(qty_edit.value),
                                    "machine": "N/A", "dt": datetime.now().strftime("%Y-%m-%d %H:%M")
                                }).execute()
                                dlg.open = False
                                page.update()
                                build_stock_list()
                            except Exception as ex:
                                page.snack_bar = ft.SnackBar(ft.Text(f"Erreur: {ex}"))
                                page.snack_bar.open = True
                                page.update()

                        dlg = ft.AlertDialog(
                            title=ft.Text(f"Modifier {item['designation']}"),
                            content=ft.Column([ft.Text(f"Stock actuel: {item['stock_qty']}"), qty_edit], height=100),
                            actions=[
                                ft.TextButton("Annuler", on_click=lambda _: (setattr(dlg, "open", False), page.update())),
                                ft.ElevatedButton("Valider", on_click=update_qty)
                            ]
                        )
                        page.dialog = dlg
                        dlg.open = True
                        page.update()

                    def open_add_part(e):
                        r_f = ft.TextField(label="Référence")
                        d_f = ft.TextField(label="Désignation")
                        c_f = ft.TextField(label="Catégorie")
                        q_f = ft.TextField(label="Stock Initial", value="0")
                        m_f = ft.TextField(label="Stock Minimum", value="1")
                        l_f = ft.TextField(label="Emplacement")

                        def save_new(e):
                            try:
                                supabase.table("inventory").insert({
                                    "ref": r_f.value, "designation": d_f.value, "category": c_f.value,
                                    "stock_qty": int(q_f.value), "min_qty": int(m_f.value), "location": l_f.value
                                }).execute()
                                add_dlg.open = False
                                page.update()
                                build_stock_list()
                            except Exception as ex:
                                page.snack_bar = ft.SnackBar(ft.Text(f"Erreur: {ex}"))
                                page.snack_bar.open = True
                                page.update()

                        add_dlg = ft.AlertDialog(
                            title=ft.Text("Nouvelle Pièce"),
                            content=ft.Column([r_f, d_f, c_f, q_f, m_f, l_f], scroll=ft.ScrollMode.AUTO),
                            actions=[ft.ElevatedButton("Ajouter", on_click=save_new)]
                        )
                        page.dialog = add_dlg
                        add_dlg.open = True
                        page.update()

                    search_stock = ft.TextField(
                        label="Chercher une pièce...",
                        prefix_icon=ft.icons.SEARCH,
                        on_change=lambda e: build_stock_list(e.control.value)
                    )

                    screen = safe_screen(
                        ft.Row([ft.IconButton(ft.icons.ARROW_BACK, on_click=lambda _: ch_v("HOME")), header_brand]),
                        ft.Row([ft.Text("GESTION DU STOCK PIÈCES", size=20, weight="bold"), ft.Icon(ft.icons.INVENTORY_2, color="teal")]),
                        search_stock,
                        ft.Row([
                            ft.ElevatedButton("AJOUTER PIÈCE", icon=ft.icons.ADD, on_click=open_add_part, bgcolor="teal"),
                            ft.ElevatedButton("EXPORT EXCEL", icon=ft.icons.FILE_DOWNLOAD, on_click=export_inventory_excel, bgcolor="green700"),
                        ]),
                        stock_lv,
                        footer_tag
                    )
                    build_stock_list()

                elif page.view == "ROUTINE":
                    m_dd = ft.Dropdown(label="Machine", options=[ft.dropdown.Option(m) for m in MACHINES])
                    c_grease = ft.Checkbox(label="Graissage")
                    c_oil = ft.Checkbox(label="Huilage")
                    c_elec = ft.Checkbox(label="Serrage électriques")
                    c_sec = ft.Checkbox(label="Test sécurité")
                    dur_in = ft.TextField(label="Temps passé (Minutes)", keyboard_type="number")

                    def save_r(e):
                        try:
                            supabase.table("routines").insert({
                                "machine": m_dd.value, "freq": "Daily", "graissage": str(c_grease.value),
                                "huilage": str(c_oil.value), "serrage": str(c_elec.value),
                                "securite": str(c_sec.value), "duree": dur_in.value,
                                "dt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "user": page.display_name
                            }).execute()
                            ch_v("HOME")
                        except Exception as ex:
                            page.snack_bar = ft.SnackBar(ft.Text(f"Erreur: {ex}"))
                            page.snack_bar.open = True
                            page.update()

                    screen = safe_screen(
                        ft.Row([ft.IconButton(ft.icons.ARROW_BACK, on_click=lambda _: ch_v("HOME")), header_brand]),
                        ft.Text("CHECK QUOTIDIEN", weight="bold"),
                        m_dd, c_grease, c_oil, c_elec, c_sec, dur_in,
                        ft.ElevatedButton("SAUVEGARDER", on_click=save_r, bgcolor="green", width=320),
                        ft.Row([
                            ft.ElevatedButton("HISTORIQUE", icon=ft.icons.HISTORY, on_click=lambda _: ch_v("ROUTINE_HISTORY")),
                            ft.ElevatedButton("RAPPORT HEBDO (PDF)", icon=ft.icons.SUMMARIZE, on_click=generate_weekly_pdf, bgcolor="blue900")
                        ]),
                        footer_tag
                    )

                elif page.view == "ROUTINE_HISTORY":
                    rows = supabase.table("routines").select("*").order("id", desc=True).execute().data
                    lv = ft.ListView(expand=True)

                    for r in rows:
                        lv.controls.append(ft.Container(
                            content=ft.Column([
                                ft.Text(f"{r['dt']} - {r['machine']}", weight="bold"),
                                ft.Text(f"Par: {r['user']} | Durée: {r['duree']} min"),
                                ft.Text(f"Checklist: G:{r['graissage']} H:{r['huilage']} S:{r['serrage']} Sec:{r['securite']}", size=12)
                            ]),
                            padding=10,
                            border=ft.border.all(1, "grey700"),
                            border_radius=10
                        ))

                    screen = safe_screen(
                        ft.Row([ft.IconButton(ft.icons.ARROW_BACK, on_click=lambda _: ch_v("ROUTINE")), header_brand]),
                        ft.Text("HISTORIQUE DES INSPECTIONS", weight="bold"),
                        ft.ElevatedButton("EXPORTER EXCEL", icon=ft.icons.FILE_DOWNLOAD, on_click=export_routines_excel, bgcolor="green700"),
                        lv,
                        footer_tag
                    )

                elif page.view == "INTER":
                    dem = ft.TextField(label="Demandeur", value="Production")
                    mold_ref_field = ft.TextField(label="Référence du Moule", visible=False)

                    def on_sys_change(e):
                        mold_ref_field.visible = (sys_dd.value == "Moule")
                        page.update()

                    sys_dd = ft.Dropdown(label="Machine / Système", options=[ft.dropdown.Option(m) for m in MACHINES], on_change=on_sys_change)
                    s_ens = ft.TextField(label="Sous-Ensemble")
                    m_type = ft.Dropdown(label="Type Maintenance", options=[ft.dropdown.Option("Corrective"), ft.dropdown.Option("Préventive")])
                    piec = ft.TextField(label="Pièces de rechange")
                    spare_price_in = ft.TextField(label="Coût Total Pièces (DZD)", keyboard_type="number", value="0")

                    error_other_desc = ft.TextField(label="Précisez le type d'erreur", visible=False)

                    def on_source_change(e):
                        error_other_desc.visible = (error_source.value == "Autre")
                        page.update()

                    error_source = ft.Dropdown(label="Source de l'erreur", options=[
                        ft.dropdown.Option("Opérateur"),
                        ft.dropdown.Option("Technique"),
                        ft.dropdown.Option("Autre")
                    ], on_change=on_source_change)

                    err_desc = ft.TextField(label="Identification de l'Erreur", multiline=True)
                    sol_desc = ft.TextField(label="Solution Apportée", multiline=True)

                    def pick_img(target):
                        page.current_upload_target = target
                        file_picker.pick_files()

                    def save_i(e):
                        try:
                            if piec.value:
                                part_res = supabase.table("inventory").select("id, stock_qty").eq("designation", piec.value).execute()
                                if part_res.data:
                                    part = part_res.data[0]
                                    new_stock = part['stock_qty'] - 1
                                    supabase.table("inventory").update({"stock_qty": new_stock}).eq("id", part['id']).execute()
                                    supabase.table("inventory_logs").insert({
                                        "part_id": part['id'],
                                        "action": "Reserved for Maint.",
                                        "qty": -1,
                                        "machine": sys_dd.value,
                                        "dt": datetime.now().strftime("%Y-%m-%d %H:%M")
                                    }).execute()

                            final_err_img_url = upload_to_supabase(page.photo_err_path, "errors")
                            final_sol_img_url = upload_to_supabase(page.photo_sol_path, "solutions")

                            supabase.table("inters").insert({
                                "demandeur": dem.value,
                                "intervenant": page.display_name,
                                "date": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                                "systeme": sys_dd.value,
                                "sous_ens": s_ens.value,
                                "error_desc": err_desc.value,
                                "solution_desc": sol_desc.value,
                                "pieces": piec.value,
                                "type_m": m_type.value,
                                "photo_err": final_err_img_url,
                                "photo_sol": final_sol_img_url,
                                "error_source": error_source.value,
                                "error_other": error_other_desc.value,
                                "user": page.display_name,
                                "mold_ref": mold_ref_field.value,
                                "spare_price": spare_price_in.value
                            }).execute()

                            page.photo_err_path = ""
                            page.photo_sol_path = ""
                            ch_v("HOME")
                        except Exception as ex:
                            page.snack_bar = ft.SnackBar(ft.Text(f"Erreur: {ex}"))
                            page.snack_bar.open = True
                            page.update()

                    screen = safe_screen(
                        ft.Row([ft.IconButton(ft.icons.ARROW_BACK, on_click=lambda _: ch_v("HOME")), header_brand]),
                        ft.Text("NOUVEAU RAPPORT D'INTERVENTION", weight="bold"),
                        dem, sys_dd, mold_ref_field, s_ens, m_type,
                        error_source, error_other_desc,
                        ft.Row([err_desc, ft.IconButton(ft.icons.CAMERA_ALT, on_click=lambda _: pick_img("ERR"), icon_color="red")]),
                        ft.Row([sol_desc, ft.IconButton(ft.icons.CAMERA_ALT, on_click=lambda _: pick_img("SOL"), icon_color="green")]),
                        piec, spare_price_in,
                        ft.ElevatedButton("ENREGISTRER & DESTOCKER", on_click=save_i, bgcolor="blue", width=320),
                        footer_tag
                    )

                elif page.view == "HISTORY":
                    lv = ft.ListView(expand=True, spacing=10)

                    def build_history(search_term=""):
                        try:
                            lv.controls.clear()
                            query = supabase.table("inters").select("*")
                            if search_term:
                                query = query.or_(f"systeme.ilike.%{search_term}%,user.ilike.%{search_term}%")
                            reports = query.order("id", desc=True).execute().data

                            for r in reports:
                                lv.controls.append(ft.Container(content=ft.Column([
                                    ft.Row([ft.Text(f"REF: INT-2026-{r['id']} | {r['systeme']}", weight="bold")]),
                                    ft.Row([ft.IconButton(ft.icons.PICTURE_AS_PDF, icon_color="red", on_click=lambda e, row=r: export_pdf(row))])
                                ]), padding=10, border=ft.border.all(1, "grey700"), border_radius=10))
                            page.update()
                        except Exception as ex:
                            print("HISTORY ERROR:", ex)

                    search_bar = ft.TextField(label="Chercher par machine...", prefix_icon=ft.icons.SEARCH, on_change=lambda e: build_history(e.control.value))
                    screen = safe_screen(
                        ft.Row([ft.IconButton(ft.icons.ARROW_BACK, on_click=lambda _: ch_v("HOME")), header_brand]),
                        ft.Text("HISTORIQUE DES INTERVENTIONS", size=20, weight="bold"),
                        search_bar,
                        ft.ElevatedButton("EXPORTER EXCEL (INTERVENTIONS)", icon=ft.icons.TABLE_CHART, on_click=export_excel, width=320),
                        lv,
                        footer_tag
                    )
                    build_history()

                elif page.view == "MOLD":
                    p_dd = ft.Dropdown(label="Presse", options=[ft.dropdown.Option(f"Presse {i+1}") for i in range(11)])
                    o_m = ft.TextField(label="Moule Sortant", read_only=True)
                    n_m = ft.TextField(label="Nouveau Moule")

                    def on_p_change(e):
                        try:
                            res = supabase.table("molds").select("new_m").eq("p_no", p_dd.value).order("id", desc=True).limit(1).execute()
                            o_m.value = res.data[0]['new_m'] if res.data else "Vide"
                            page.update()
                        except Exception as ex:
                            print("MOLD LOAD ERROR:", ex)

                    p_dd.on_change = on_p_change

                    def save_m(e):
                        try:
                            supabase.table("molds").insert({
                                "p_no": p_dd.value,
                                "old_m": o_m.value,
                                "new_m": n_m.value,
                                "dt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                "user": page.display_name
                            }).execute()
                            ch_v("HOME")
                        except Exception as ex:
                            page.snack_bar = ft.SnackBar(ft.Text(f"Erreur: {ex}"))
                            page.snack_bar.open = True
                            page.update()

                    screen = safe_screen(
                        ft.Row([ft.IconButton(ft.icons.ARROW_BACK, on_click=lambda _: ch_v("HOME")), header_brand]),
                        ft.Text("CHANGEMENT MOULE (SMED)", weight="bold"),
                        p_dd, o_m, n_m,
                        ft.ElevatedButton("VALIDER", on_click=save_m, width=320),
                        ft.ElevatedButton("VOIR HISTORIQUE", icon=ft.icons.HISTORY, on_click=lambda _: ch_v("MOLD_HISTORY"), width=320, bgcolor="blue900"),
                        ft.ElevatedButton("EXPORTER EXCEL (MOULES)", icon=ft.icons.TABLE_CHART, on_click=export_molds_excel, width=320, bgcolor="green700"),
                        footer_tag
                    )

                elif page.view == "MOLD_HISTORY":
                    m_reports = supabase.table("molds").select("*").order("id", desc=True).execute().data
                    lv = ft.ListView(expand=True, spacing=10)

                    for r in m_reports:
                        lv.controls.append(ft.Container(content=ft.Column([
                            ft.Row([ft.Text(f"{r['dt']} - {r['p_no']}", weight="bold")]),
                            ft.Row([ft.Text(f"Sortant: {r['old_m']}   ➔ Nouveau: {r['new_m']}")])
                        ]), padding=10, border=ft.border.all(1, "grey700"), border_radius=10))

                    screen = safe_screen(
                        ft.Row([ft.IconButton(ft.icons.ARROW_BACK, on_click=lambda _: ch_v("MOLD")), header_brand]),
                        ft.Text("HISTORIQUE DES MOULES", weight="bold"),
                        lv,
                        footer_tag
                    )

                else:
                    screen = safe_screen(
                        header_brand,
                        ft.Text(f"Vue inconnue: {page.view}", color="red"),
                        ft.ElevatedButton("Retour Accueil", on_click=lambda _: ch_v("HOME")),
                        footer_tag
                    )

                content_area.content = screen

                if content_area not in page.controls:
                    page.controls.clear()
                    page.add(content_area)

                page.update()

            except Exception as e:
                print(f"Error in refresh: {e}")
                content_area.content = safe_screen(
                    ft.Text("ERREUR D'AFFICHAGE", color="red", size=24, weight="bold"),
                    ft.Text(str(e), color="white"),
                    ft.ElevatedButton("Retour Login", on_click=lambda _: ch_v("LOGIN")),
                    footer_tag
                )
                if content_area not in page.controls:
                    page.controls.clear()
                    page.add(content_area)
                page.update()

        def export_routines_excel(e=None):
            try:
                data = supabase.table("routines").select("*").execute().data
                df = pd.DataFrame(data)
                base_name = f"Briks_By_Okba_Daily_Inspections_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                filename = os.path.join(tempfile.gettempdir(), base_name)
                df.to_excel(filename, index=False)
                page.snack_bar = ft.SnackBar(ft.Text(f"Exporté : {base_name}"))
                page.snack_bar.open = True
                page.update()
            except Exception as ex:
                page.snack_bar = ft.SnackBar(ft.Text(f"Erreur export: {ex}"))
                page.snack_bar.open = True
                page.update()

        def export_excel(e=None):
            try:
                data = supabase.table("inters").select("*").execute().data
                df = pd.DataFrame(data)
                base_name = f"Briks_By_Okba_Rapports_Global_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                filename = os.path.join(tempfile.gettempdir(), base_name)
                df.to_excel(filename, index=False)
                page.snack_bar = ft.SnackBar(ft.Text(f"Excel exporté : {base_name}"))
                page.snack_bar.open = True
                page.update()
            except Exception as ex:
                page.snack_bar = ft.SnackBar(ft.Text(f"Erreur export: {ex}"))
                page.snack_bar.open = True
                page.update()

        def export_molds_excel(e=None):
            try:
                data = supabase.table("molds").select("*").execute().data
                df = pd.DataFrame(data)
                base_name = f"Briks_By_Okba_Tracking_Moules_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                filename = os.path.join(tempfile.gettempdir(), base_name)
                df.to_excel(filename, index=False)
                page.snack_bar = ft.SnackBar(ft.Text(f"Excel exporté : {base_name}"))
                page.snack_bar.open = True
                page.update()
            except Exception as ex:
                page.snack_bar = ft.SnackBar(ft.Text(f"Erreur export: {ex}"))
                page.snack_bar.open = True
                page.update()

        def export_inventory_excel(e=None):
            try:
                data = supabase.table("inventory").select("*").execute().data
                df = pd.DataFrame(data)
                base_name = f"Briks_By_Okba_Stock_Inventory_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                filename = os.path.join(tempfile.gettempdir(), base_name)
                df.to_excel(filename, index=False)
                page.snack_bar = ft.SnackBar(ft.Text(f"Inventaire exporté : {base_name}"))
                page.snack_bar.open = True
                page.update()
            except Exception as ex:
                page.snack_bar = ft.SnackBar(ft.Text(f"Erreur export: {ex}"))
                page.snack_bar.open = True
                page.update()

        def generate_weekly_pdf(e=None):
            try:
                last_week = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
                data = supabase.table("routines").select("*").gte("dt", last_week).execute().data

                pdf = FPDF()
                pdf.add_page()
                pdf.set_font("Arial", 'B', 16)
                pdf.cell(190, 10, "BRIKS BY OKBA - RAPPORT HEBDOMADAIRE", ln=True, align='C')
                pdf.set_font("Arial", '', 10)
                pdf.cell(190, 10, f"Période: du {last_week} au {datetime.now().strftime('%Y-%m-%d')}", ln=True, align='C')
                pdf.ln(10)

                pdf.set_font("Arial", 'B', 10)
                pdf.cell(40, 10, "Date/Heure", 1)
                pdf.cell(40, 10, "Machine", 1)
                pdf.cell(30, 10, "Opérateur", 1)
                pdf.cell(80, 10, "Points de contrôle", 1)
                pdf.ln()

                pdf.set_font("Arial", '', 8)
                for r in data:
                    checklist = f"Gr:{r['graissage']}, Hu:{r['huilage']}, Se:{r['serrage']}, Sec:{r['securite']}"
                    pdf.cell(40, 8, str(r['dt']), 1)
                    pdf.cell(40, 8, str(r['machine']), 1)
                    pdf.cell(30, 8, str(r['user']), 1)
                    pdf.cell(80, 8, checklist, 1)
                    pdf.ln()

                base_name = f"Rapport_Hebdo_{datetime.now().strftime('%Y%W')}.pdf"
                fname = os.path.join(tempfile.gettempdir(), base_name)
                pdf.output(fname)

                page.snack_bar = ft.SnackBar(ft.Text(f"Rapport Hebdo créé: {base_name}"))
                page.snack_bar.open = True
                page.update()
            except Exception as ex:
                page.snack_bar = ft.SnackBar(ft.Text(f"Erreur : {ex}"))
                page.snack_bar.open = True
                page.update()

        def export_pdf(row):
            try:
                pdf = FPDF()
                pdf.add_page()

                pdf.set_font("Arial", 'B', 22)
                pdf.set_text_color(140, 0, 0)
                pdf.cell(190, 10, "BRIKS BY OKBA", ln=True, align='C')
                pdf.set_font("Arial", 'I', 10)
                pdf.cell(190, 6, "RAPPORT D'INTERVENTION TECHNIQUE", ln=True, align='C')
                pdf.ln(5)

                pdf.set_text_color(0, 0, 0)
                pdf.set_font("Arial", 'B', 9)

                pdf.cell(100, 8, f"Machine: {row.get('systeme', '')}", border=1)
                pdf.cell(90, 8, f"Date: {row.get('date', '')}", border=1, ln=True)
                pdf.cell(100, 8, f"Source de l'erreur: {row.get('error_source', '')}", border=1)
                pdf.cell(90, 8, f"Intervenant: {row.get('intervenant', '')}", border=1, ln=True)
                pdf.ln(5)

                def draw_section(title, desc, photo_url, y_start):
                    pdf.set_xy(10, y_start)
                    pdf.set_font("Arial", 'B', 9)
                    pdf.cell(95, 8, title, border='LTR', ln=True)
                    pdf.set_font("Arial", '', 8)

                    pdf.multi_cell(95, 5, str(desc), border='LR')
                    current_y = pdf.get_y()
                    if current_y < y_start + 48:
                        pdf.cell(95, (y_start + 48) - current_y, "", border='LRB', ln=True)
                    else:
                        pdf.cell(95, 0, "", border='T', ln=True)

                    pdf.set_xy(110, y_start)
                    pdf.set_font("Arial", 'B', 9)
                    pdf.cell(80, 8, "PHOTO:", border=1, ln=True)
                    pdf.set_xy(110, y_start + 8)
                    pdf.cell(80, 40, "", border=1, ln=True)

                    if photo_url and str(photo_url).startswith("http"):
                        try:
                            temp_img = os.path.join(tempfile.gettempdir(), f"temp_img_{os.urandom(4).hex()}.jpg")
                            urllib.request.urlretrieve(photo_url, temp_img)
                            pdf.image(temp_img, x=112, y=y_start + 10, w=76, h=36)
                            os.remove(temp_img)
                        except:
                            pass

                    return max(pdf.get_y(), y_start + 48) + 5

                y_next = draw_section("1. IDENTIFICATION DE L'ERREUR:", row.get('error_desc', ''), row.get('photo_err', ''), pdf.get_y())
                y_next = draw_section("2. SOLUTION APPORTEE:", row.get('solution_desc', ''), row.get('photo_sol', ''), y_next)

                pdf.set_xy(10, y_next)
                pdf.set_font("Arial", 'B', 9)
                pdf.cell(190, 8, "CONSOMMABLES & PIECES DE RECHANGE:", border=1, ln=True)
                pdf.set_font("Arial", '', 8)
                pdf.cell(190, 8, f"Description: {row.get('pieces', '')}", border='LR', ln=True)
                pdf.cell(190, 8, f"Coût: {row.get('spare_price', '0')} DZD", border='LRB', ln=True)

                pdf.ln(20)
                pdf.set_font("Arial", '', 9)
                pdf.cell(95, 5, "Chef de Production", align='C')
                pdf.cell(95, 5, "Service Maintenance", align='C', ln=True)

                base_name = f"Rapport_{row.get('id', 'N')}_{datetime.now().strftime('%H%M%S')}.pdf"
                fname = os.path.join(tempfile.gettempdir(), base_name)
                pdf.output(fname)

                page.snack_bar = ft.SnackBar(ft.Text(f"Rapport PDF généré : {base_name}"))
                page.snack_bar.open = True
                page.update()
            except Exception as ex:
                page.snack_bar = ft.SnackBar(ft.Text(f"Erreur PDF: {ex}"))
                page.snack_bar.open = True
                page.update()

        def export_part_pdf(row):
            try:
                pdf = FPDF()
                pdf.add_page()
                pdf.set_font("Arial", 'B', 16)
                pdf.cell(190, 10, "DEMANDE DE PIECE DE RECHANGE", ln=True, align='C')
                pdf.ln(10)

                pdf.set_font("Arial", '', 12)
                pdf.cell(190, 8, f"Reference: PR-2026-{row.get('id', '')}", ln=True)
                pdf.cell(190, 8, f"Piece: {row.get('piece_nom', '')}", ln=True)
                pdf.cell(190, 8, f"Machine: {row.get('machine', '')}", ln=True)
                pdf.cell(190, 8, f"Quantite: {row.get('qte', '')}", ln=True)
                pdf.cell(190, 8, f"Urgence: {row.get('urgence', '')}", ln=True)

                if row.get('photo_path') and str(row.get('photo_path')).startswith("http"):
                    try:
                        temp_img = os.path.join(tempfile.gettempdir(), f"temp_part_{os.urandom(4).hex()}.jpg")
                        urllib.request.urlretrieve(row['photo_path'], temp_img)
                        pdf.image(temp_img, x=10, y=pdf.get_y() + 10, w=100)
                        os.remove(temp_img)
                    except:
                        pass

                base_name = f"Demande_Piece_{row.get('id', 'N')}.pdf"
                fname = os.path.join(tempfile.gettempdir(), base_name)
                pdf.output(fname)

                page.snack_bar = ft.SnackBar(ft.Text(f"PDF généré : {base_name}"))
                page.snack_bar.open = True
                page.update()
            except Exception as ex:
                page.snack_bar = ft.SnackBar(ft.Text(f"Erreur PDF: {ex}"))
                page.snack_bar.open = True
                page.update()

        refresh()

    except Exception as fatal_error:
        page.add(
            ft.Text(
                f"CRITICAL ERROR: {str(fatal_error)}",
                color="white",
                bgcolor="red",
                size=20,
                weight="bold"
            )
        )
        page.update()

ft.app(target=main)